from openpyxl import (
    Workbook,
    load_workbook,
)

from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)

from openpyxl.comments import (
    Comment
)

from form_layout import (
    MANUAL_REVIEW_THRESHOLD
)


YELLOW = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

GREEN = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9"
)

THIN = Side(
    style="thin",
    color="777777"
)


COLUMN_ORDER = [

    "floor",
    "space",
    "light_fc",
    "temp_f",
    "rh_pct",
    "time",
    "co2_ppm",
    "voc_ug_m3",
    "comments",
]


HEADERS = {

    "floor": "Floor",

    "space": "Space",

    "light_fc": "Light (fc)",

    "temp_f": "Temp (°F)",

    "rh_pct": "RH (%)",

    "time": "Time",

    "co2_ppm": "CO2 (ppm)",

    "voc_ug_m3": "VOC (µg/m3)",

    "comments": "COMMENTS",
}


def build_workbook(
    results,
    output_path,
    template_path=None
):

    if template_path:

        workbook = (
            load_workbook(
                template_path
            )
        )

    else:

        workbook = Workbook()

    worksheet = (
        workbook.active
    )

    worksheet.title = (
        "Transcribed Form"
    )


    # -------------------------
    # CREATE BASIC LAYOUT
    # -------------------------

    if not template_path:

        worksheet["A1"] = (
            "Building Name:"
        )

        worksheet["A2"] = (
            "Completed By:"
        )

        worksheet["A3"] = (
            "Date:"
        )

        worksheet["A4"] = (
            "Building Escort:"
        )

        worksheet["A5"] = (
            "CO2 Meter #:"
        )

        for column, field in enumerate(
            COLUMN_ORDER,
            start=1
        ):

            cell = worksheet.cell(
                row=7,
                column=column,
                value=HEADERS[field]
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = (
                Alignment(
                    horizontal="center",
                    wrap_text=True
                )
            )

            cell.border = (
                Border(
                    left=THIN,
                    right=THIN,
                    top=THIN,
                    bottom=THIN
                )
            )


    # -------------------------
    # HEADER DATA
    # -------------------------

    header_map = {

        "building_name": "B1",

        "completed_by": "B2",

        "date": "B3",

        "building_escort": "B4",

        "co2_meter_number": "B5",
    }


    row_data = {}


    for item in results:

        field = item["field"]

        if (
            item["row"] is None
            and
            field in header_map
        ):

            cell = worksheet[
                header_map[field]
            ]

            cell.value = item["text"]

            if (
                item["confidence"]
                <
                MANUAL_REVIEW_THRESHOLD
            ):

                cell.fill = YELLOW

                cell.comment = Comment(
                    (
                        "Needs review: "
                        f"{item['confidence']:.1%}"
                    ),
                    "Form Transcriber"
                )

        elif item["row"] is not None:

            row_data.setdefault(
                item["row"],
                {}
            )[field] = item


    # -------------------------
    # TABLE DATA
    # -------------------------

    for source_row, fields in sorted(
        row_data.items()
    ):

        excel_row = (
            8
            +
            source_row
        )

        for column, field in enumerate(
            COLUMN_ORDER,
            start=1
        ):

            item = fields.get(
                field
            )

            if not item:
                continue

            cell = worksheet.cell(
                row=excel_row,
                column=column,
                value=item["text"]
            )

            cell.border = Border(
                left=THIN,
                right=THIN,
                top=THIN,
                bottom=THIN
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if (
                item["confidence"]
                <
                MANUAL_REVIEW_THRESHOLD
            ):

                cell.fill = YELLOW

                cell.comment = Comment(
                    (
                        "Needs review: "
                        f"{item['confidence']:.1%}"
                    ),
                    "Form Transcriber"
                )


    # -------------------------
    # REVIEW SHEET
    # -------------------------

    review_sheet = (
        workbook.create_sheet(
            "Manual Review"
        )
    )

    review_sheet.append(
        [
            "Page",
            "Field",
            "Source Row",
            "Recognized Text",
            "Confidence",
            "Status",
        ]
    )

    for cell in review_sheet[1]:

        cell.font = Font(
            bold=True
        )


    for item in results:

        if (
            item["confidence"]
            <
            MANUAL_REVIEW_THRESHOLD
        ):

            review_sheet.append(
                [
                    item["page"],
                    item["field"],
                    item["row"],
                    item["text"],
                    item["confidence"],
                    "REVIEW",
                ]
            )


    workbook.save(
        output_path
    )