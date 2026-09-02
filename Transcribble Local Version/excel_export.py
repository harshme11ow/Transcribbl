from openpyxl import (
    Workbook,
    load_workbook,
)

from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)

from openpyxl.comments import (
    Comment
)


YELLOW = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC"
)

GREEN = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9"
)

# Added a light gray fill for headers to mimic a professional form
GRAY = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2"
)

THIN = Side(
    style="thin",
    color="777777"
)


COLUMN_ORDER = [
    "floor",
    "space",
    "light_fc",
    "temp_f",
    "rh_pct",
    "time",
    "co2_ppm",
    "voc_ug_m3",
    "comments",
]


def build_workbook(
    results,
    output_path,
    template_path=None
):

    # Fallback in case the new VLM setup bypasses form_layout imports
    try:
        from form_layout import MANUAL_REVIEW_THRESHOLD
    except ImportError:
        MANUAL_REVIEW_THRESHOLD = 0.82

    if template_path:
        workbook = load_workbook(template_path)
    else:
        workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Transcribed Form"


    # -------------------------
    # CREATE IDENTICAL LAYOUT
    # -------------------------

    if not template_path:
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        bottom_line = Border(bottom=THIN)

        # 1. Setup exact column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 12
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 16
        worksheet.column_dimensions['I'].width = 45

        # 2. Add Instruction Checkboxes
        instructions = [
            "☐ Inspect major ventilation systems and take screenshots of BAS. If there are AHUs per floor, inspect a sample of them.",
            "☐ Find all the energy and water meters. Take a picture and confirm what area they serve.",
            "☐ Take at least one reading per floor in an occupied tenant space. If vacant, unoccupied, or under construction, note below.",
            "☐ Take a picture of the devices after each reading and mark the location of each reading on the floor plans.",
            "☐ Message LEED Project Manager to end TVOC rental with Pine."
        ]
        for i, text in enumerate(instructions, start=1):
            worksheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
            cell = worksheet.cell(row=i, column=1, value=text)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            worksheet.row_dimensions[i].height = 25

        # 3. Add Header Fields
        headers_list = ["Building Name:", "Completed By:", "Date:", "Building Escort:", "CO2 Meter #:"]
        for i, label in enumerate(headers_list, start=1):
            cell = worksheet.cell(row=i, column=6, value=label)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Draw the underline for the field value
            worksheet.merge_cells(start_row=i, start_column=7, end_row=i, end_column=9)
            worksheet.cell(row=i, column=7).border = bottom_line
            worksheet.cell(row=i, column=8).border = bottom_line
            worksheet.cell(row=i, column=9).border = bottom_line
            
            worksheet.cell(row=i, column=7).alignment = Alignment(horizontal="left", vertical="bottom")

        # 4. Add Table Headers (Rows 7 and 8)
        worksheet.row_dimensions[7].height = 25
        worksheet.row_dimensions[8].height = 40

        worksheet.merge_cells("A7:A8")
        worksheet.merge_cells("B7:B8")
        worksheet.merge_cells("C7:E7")
        worksheet.merge_cells("F7:H7")
        worksheet.merge_cells("I7:I8")

        worksheet["A7"] = "Floor"
        worksheet["B7"] = "Space\n(tenant, suite number + open office, private office, or reception)"
        worksheet["C7"] = "IAQ"
        worksheet["F7"] = "Data Point 1"
        worksheet["I7"] = "COMMENTS\n(approx. # of people, any smells, damper closed/open, notes about tenant)"

        worksheet["C8"] = "Light (fc)"
        worksheet["D8"] = "Temp (°F)"
        worksheet["E8"] = "RH (%)"
        worksheet["F8"] = "Time"
        worksheet["G8"] = "CO2 (ppm)"
        worksheet["H8"] = "VOC (µg/m3)"

        # Apply formatting to all table header cells
        for row in [7, 8]:
            for col in range(1, 10):
                c = worksheet.cell(row=row, column=col)
                c.font = bold_font
                c.fill = GRAY
                c.border = thin_border
                c.alignment = center_aligned


    # -------------------------
    # WRITE HEADER DATA
    # -------------------------

    # The new visual layout places these fields starting at G1 (column 7)
    header_map = {
        "building_name": "G1",
        "completed_by": "G2",
        "date": "G3",
        "building_escort": "G4",
        "co2_meter_number": "G5",
    }

    row_data = {}

    for item in results:
        field = item["field"]

        if item["row"] is None and field in header_map:
            cell = worksheet[header_map[field]]
            cell.value = item["text"]

            if item["confidence"] < MANUAL_REVIEW_THRESHOLD:
                cell.fill = YELLOW
                cell.comment = Comment(
                    f"Needs review: {item['confidence']:.1%}",
                    "Form Transcriber"
                )

        elif item["row"] is not None:
            row_data.setdefault(item["row"], {})[field] = item


    # -------------------------
    # WRITE TABLE DATA
    # -------------------------

    # Data rows now start below the double-header at row 9
    for source_row, fields in sorted(row_data.items()):
        try:
            row_idx = int(source_row)
        except ValueError:
            continue
            
        excel_row = 9 + row_idx
        
        # Give the row breathing room to match the physical form
        worksheet.row_dimensions[excel_row].height = 25

        for column, field in enumerate(COLUMN_ORDER, start=1):
            item = fields.get(field)

            if not item:
                # Draw empty borders for unpopulated grid cells
                empty_cell = worksheet.cell(row=excel_row, column=column)
                empty_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                continue

            cell = worksheet.cell(
                row=excel_row,
                column=column,
                value=item["text"]
            )

            cell.border = Border(
                left=THIN,
                right=THIN,
                top=THIN,
                bottom=THIN
            )

            # Left align long text like comments and spaces; center numbers
            if field in ["comments", "space"]:
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

            if item["confidence"] < MANUAL_REVIEW_THRESHOLD:
                cell.fill = YELLOW
                cell.comment = Comment(
                    f"Needs review: {item['confidence']:.1%}",
                    "Form Transcriber"
                )


    # -------------------------
    # REVIEW SHEET
    # -------------------------

    review_sheet = workbook.create_sheet("Manual Review")
    review_sheet.append([
        "Page",
        "Field",
        "Source Row",
        "Recognized Text",
        "Confidence",
        "Status",
    ])

    for cell in review_sheet[1]:
        cell.font = Font(bold=True)

    for item in results:
        if item["confidence"] < MANUAL_REVIEW_THRESHOLD:
            review_sheet.append([
                item.get("page", ""),
                item["field"],
                item.get("row", ""),
                item["text"],
                item["confidence"],
                "REVIEW",
            ])

    workbook.save(output_path)